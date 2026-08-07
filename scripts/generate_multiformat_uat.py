from __future__ import annotations

import csv
import io
import textwrap
import zipfile
from html import escape
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageDraw, ImageFont
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


ROOT = Path(__file__).parents[1]
OUTPUT = ROOT / "docs" / "uat" / "input_formats"
TITLE = "Synthetic Annual Microsoft Licensing Requirement"
CUSTOMER = "Northwind Enterprise Demo (Synthetic)"
HEADERS = [
    "ProductId",
    "SkuId",
    "SKU",
    "Quantity",
    "Subscription Term",
    "Billing Plan",
    "Renewal Date",
]
ROWS = [
    ["CFQ7TTC0LFLX", "1", "Microsoft 365 E3", 120, "P1Y", "Annual", "2026-09-30"],
    ["CFQ7TTC0LHSF", "1", "Power BI Pro", 30, "P1Y", "Annual", "2026-10-15"],
    [
        "CFQ7TTC0LHT4",
        "1",
        "Enterprise Mobility + Security E3",
        120,
        "P1Y",
        "Annual",
        "2026-09-30",
    ],
    [
        "CFQ7TTC0LH0T",
        "1",
        "Microsoft Teams Phone Standard",
        60,
        "P1Y",
        "Annual",
        "2026-12-31",
    ],
    [
        "CFQ7TTC0LH04",
        "1",
        "Microsoft Defender for Office 365 (Plan 1)",
        120,
        "P1Y",
        "Annual",
        "2026-09-30",
    ],
]


def main() -> None:
    OUTPUT.mkdir(parents=True, exist_ok=True)
    _write_delimited(OUTPUT / "licensing_requirement.csv", ",")
    _write_delimited(OUTPUT / "licensing_requirement.tsv", "\t")
    _write_excel(OUTPUT / "licensing_requirement.xlsx", arbitrary_layout=False)
    _write_excel(OUTPUT / "licensing_requirement_arbitrary_layout.xlsx", arbitrary_layout=True)
    _write_text(OUTPUT / "licensing_requirement.txt")
    _write_docx(OUTPUT / "licensing_requirement.docx")
    _write_text_pdf(OUTPUT / "licensing_requirement_text.pdf")

    image = _requirement_image()
    image.save(OUTPUT / "licensing_requirement.png", format="PNG", optimize=True)
    image.convert("RGB").save(
        OUTPUT / "licensing_requirement.jpg", format="JPEG", quality=95, optimize=True
    )
    image.save(OUTPUT / "licensing_requirement.webp", format="WEBP", quality=95)
    _write_scanned_pdf(OUTPUT / "licensing_requirement_scanned.pdf", image)

    unclear = _requirement_image(unclear_quantity=True)
    unclear.convert("RGB").save(
        OUTPUT / "negative_unclear_quantity.jpg",
        format="JPEG",
        quality=95,
        optimize=True,
    )
    _write_voice_script(OUTPUT / "voice_note_script.txt")
    _write_bundle(OUTPUT.parent / "ssp_multiformat_uat_pack.zip")

    print(f"Generated {len(list(OUTPUT.iterdir()))} UAT files in {OUTPUT}")


def _write_delimited(path: Path, delimiter: str) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle, delimiter=delimiter)
        writer.writerow(HEADERS)
        writer.writerows(ROWS)


def _write_excel(path: Path, *, arbitrary_layout: bool) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Licensing Requirement"
    if arbitrary_layout:
        sheet.merge_cells("A1:G1")
        sheet["A1"] = TITLE
        sheet["A2"] = f"Customer: {CUSTOMER}"
        sheet["A3"] = "Commercial basis: one-year subscription with annual billing"
        sheet["A4"] = "Prepared for controlled UAT only"
        header_row = 6
    else:
        header_row = 1

    for column, value in enumerate(HEADERS, start=1):
        cell = sheet.cell(header_row, column, value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_number, values in enumerate(ROWS, start=header_row + 1):
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_number, column, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if row_number % 2 == 0:
            for cell in sheet[row_number]:
                cell.fill = PatternFill("solid", fgColor="F2F6FC")

    widths = [20, 10, 48, 12, 20, 16, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:G{header_row + len(ROWS)}"
    sheet.sheet_view.showGridLines = False
    workbook.save(path)


def _write_text(path: Path) -> None:
    lines = [
        TITLE,
        f"Customer: {CUSTOMER}",
        "Commercial basis: one-year subscription (P1Y), billed annually.",
        "",
        "Required licences:",
    ]
    for index, row in enumerate(ROWS, start=1):
        lines.append(
            f"{index}. {row[2]} - {row[3]} licences - {row[4]} - {row[5]} billing - "
            f"renewal {row[6]} - ProductId {row[0]} - SkuId {row[1]}"
        )
    lines.extend(
        [
            "",
            "Please capture the requirement exactly as stated and present it for seller "
            "confirmation before calculating the annual value.",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_docx(path: Path) -> None:
    def paragraph(text: str, *, bold: bool = False, size: int = 22) -> str:
        properties = f"<w:rPr>{'<w:b/>' if bold else ''}<w:sz w:val=\"{size}\"/></w:rPr>"
        return (
            "<w:p><w:r>"
            + properties
            + f'<w:t xml:space="preserve">{escape(text)}</w:t>'
            + "</w:r></w:p>"
        )

    def cell(value: object, *, bold: bool = False) -> str:
        return "<w:tc>" + paragraph(str(value), bold=bold, size=18) + "</w:tc>"

    table_rows = [
        "<w:tr>" + "".join(cell(value, bold=True) for value in HEADERS) + "</w:tr>"
    ]
    table_rows.extend(
        "<w:tr>" + "".join(cell(value) for value in row) + "</w:tr>" for row in ROWS
    )
    borders = "".join(
        f'<w:{side} w:val="single" w:sz="4" w:color="B8C8DC"/>'
        for side in ("top", "left", "bottom", "right", "insideH", "insideV")
    )
    table = f"<w:tbl><w:tblPr><w:tblBorders>{borders}</w:tblBorders></w:tblPr>{''.join(table_rows)}</w:tbl>"
    document = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:body>"
        + paragraph(TITLE, bold=True, size=32)
        + paragraph(f"Customer: {CUSTOMER}", size=22)
        + paragraph("Commercial basis: one-year subscription with annual billing.", size=22)
        + table
        + paragraph(
            "Please present this requirement for seller confirmation before pricing.",
            size=20,
        )
        + '<w:sectPr><w:pgSz w:w="16838" w:h="11906" w:orient="landscape"/></w:sectPr>'
        + "</w:body></w:document>"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        "</Types>"
    )
    relationships = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="word/document.xml"/>'
        "</Relationships>"
    )
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types)
        archive.writestr("_rels/.rels", relationships)
        archive.writestr("word/document.xml", document)


def _write_text_pdf(path: Path) -> None:
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "UATTitle",
        parent=styles["Title"],
        alignment=TA_CENTER,
        textColor=colors.HexColor("#17365D"),
    )
    document = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=TITLE,
    )
    data: list[list[object]] = [[Paragraph(escape(value), styles["BodyText"]) for value in HEADERS]]
    for row in ROWS:
        data.append([Paragraph(escape(str(value)), styles["BodyText"]) for value in row])
    table = Table(
        data,
        colWidths=[32 * mm, 12 * mm, 72 * mm, 22 * mm, 31 * mm, 28 * mm, 31 * mm],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17365D")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B8C8DC")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FC")]),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    document.build(
        [
            Paragraph(TITLE, title_style),
            Paragraph(f"Customer: {CUSTOMER}", styles["Normal"]),
            Paragraph(
                "Commercial basis: one-year subscription with annual billing.",
                styles["Normal"],
            ),
            Spacer(1, 5 * mm),
            table,
            Spacer(1, 5 * mm),
            Paragraph(
                "Please present this requirement for seller confirmation before pricing.",
                styles["Normal"],
            ),
        ]
    )


def _font(size: int, *, bold: bool = False) -> ImageFont.ImageFont:
    names = ["arialbd.ttf", "DejaVuSans-Bold.ttf"] if bold else ["arial.ttf", "DejaVuSans.ttf"]
    for name in names:
        try:
            return ImageFont.truetype(name, size)
        except OSError:
            continue
    return ImageFont.load_default()


def _requirement_image(*, unclear_quantity: bool = False) -> Image.Image:
    width, height = 1700, 990
    image = Image.new("RGB", (width, height), "#F4F7FB")
    draw = ImageDraw.Draw(image)
    title_font = _font(42, bold=True)
    subtitle_font = _font(24)
    header_font = _font(21, bold=True)
    body_font = _font(23)
    note_font = _font(22, bold=True)
    draw.rectangle((0, 0, width, 150), fill="#17365D")
    draw.text((50, 30), TITLE, font=title_font, fill="white")
    draw.text(
        (50, 92),
        f"{CUSTOMER} | One-year subscription | Annual billing",
        font=subtitle_font,
        fill="#DDE8F8",
    )

    columns = [230, 100, 560, 100, 170, 170, 270]
    labels = ["ProductId", "SkuId", "SKU", "Qty", "Term", "Billing", "Renewal date"]
    x0, y0, header_height, row_height = 50, 190, 68, 116
    x = x0
    for label, column_width in zip(labels, columns, strict=True):
        draw.rectangle((x, y0, x + column_width, y0 + header_height), fill="#1F4776", outline="#C9D5E5", width=2)
        draw.text((x + 12, y0 + 20), label, font=header_font, fill="white")
        x += column_width

    for index, row in enumerate(ROWS):
        values = list(row)
        if unclear_quantity and index == 1:
            values[3] = "?"
        top = y0 + header_height + index * row_height
        x = x0
        fill = "#FFFFFF" if index % 2 == 0 else "#EDF3FA"
        for column_index, (value, column_width) in enumerate(zip(values, columns, strict=True)):
            draw.rectangle((x, top, x + column_width, top + row_height), fill=fill, outline="#C9D5E5", width=2)
            lines = textwrap.wrap(str(value), width=34 if column_index == 2 else 22) or [""]
            text_y = top + 15
            for line in lines[:3]:
                draw.text((x + 12, text_y), line, font=body_font, fill="#172033")
                text_y += 30
            x += column_width

    note_top = y0 + header_height + len(ROWS) * row_height + 34
    if unclear_quantity:
        note = "NEGATIVE TEST: Power BI Pro quantity is intentionally unclear. Ask the seller to confirm it."
        fill, outline, text_color = "#FFF3D6", "#D98E04", "#7A4800"
    else:
        note = "Seller confirmation required before annual pricing is calculated."
        fill, outline, text_color = "#E5F5EC", "#79BF9B", "#18794E"
    draw.rounded_rectangle((50, note_top, width - 50, note_top + 82), radius=16, fill=fill, outline=outline, width=3)
    draw.text((75, note_top + 24), note, font=note_font, fill=text_color)
    return image


def _write_scanned_pdf(path: Path, image: Image.Image) -> None:
    image_buffer = io.BytesIO()
    image.save(image_buffer, format="PNG")
    image_buffer.seek(0)
    page_width, page_height = landscape(A4)
    document = canvas.Canvas(str(path), pagesize=(page_width, page_height), pageCompression=1)
    document.setTitle(TITLE)
    document.drawImage(
        ImageReader(image_buffer),
        8 * mm,
        8 * mm,
        width=page_width - 16 * mm,
        height=page_height - 16 * mm,
        preserveAspectRatio=True,
        anchor="c",
    )
    document.showPage()
    document.save()


def _write_voice_script(path: Path) -> None:
    sku_lines = "; ".join(f"{row[3]} {row[2]} licences" for row in ROWS)
    text = (
        "VOICE NOTE UAT SCRIPT - synthetic data only\n\n"
        "Read the following aloud as one WhatsApp voice note:\n\n"
        f"Please capture this one-year licensing requirement with annual billing: {sku_lines}. "
        "Present the extracted requirement for my confirmation before calculating the cost.\n\n"
        "Expected behavior: the agent displays the transcript, extracts five SKU lines, "
        "and pauses for seller confirmation.\n"
    )
    path.write_text(text, encoding="utf-8")


def _write_bundle(path: Path) -> None:
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for asset in sorted(OUTPUT.iterdir()):
            if asset.is_file():
                archive.write(asset, arcname=asset.name)


if __name__ == "__main__":
    main()
