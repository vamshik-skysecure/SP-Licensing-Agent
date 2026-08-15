from __future__ import annotations

import csv
import io
import re
from collections.abc import Iterable, Mapping, Sequence
from datetime import UTC, date, datetime
from pathlib import Path

from .models import (
    EstateStatus,
    LicenseEstate,
    NormalizedLicenseLine,
    ParsedLicenseRow,
    SellerProvidedDetail,
)
from .rate_card import RateCardCatalog, RateCardProvider


class LicenseAnalysisError(ValueError):
    pass


_HEADER_ALIASES = {
    "producttitle": "product_title",
    "productname": "product_title",
    "sku": "product_title",
    "skuname": "product_title",
    "licensename": "product_title",
    "licencename": "product_title",
    "productid": "product_id",
    "skuid": "sku_id",
    "skupartnumber": "sku_id",
    "totallicenses": "total_licenses",
    "totalquantity": "total_licenses",
    "quantity": "total_licenses",
    "qty": "total_licenses",
    "licensequantity": "total_licenses",
    "licencequantity": "total_licenses",
    "expiredlicenses": "expired_licenses",
    "assignedlicenses": "assigned_licenses",
    "consumedlicenses": "assigned_licenses",
    "expirationdate": "expiration_date",
    "expirydate": "expiration_date",
    "renewaldate": "renewal_date",
    "termduration": "term_duration",
    "subscriptionterm": "term_duration",
    "licenceterm": "term_duration",
    "licenseterm": "term_duration",
    "billingplan": "billing_plan",
    "billingterm": "billing_plan",
    "billingfrequency": "billing_plan",
}

_REQUIRED = {
    "product_title",
    "total_licenses",
}


class LicenseAnalyzer:
    def __init__(
        self,
        rate_cards: RateCardProvider,
        *,
        match_threshold: float = 90.0,
        default_term_duration: str = "P1Y",
        default_billing_plan: str = "Annual",
    ) -> None:
        self._rate_cards = rate_cards
        self._match_threshold = match_threshold
        self._default_term_duration = default_term_duration
        self._default_billing_plan = default_billing_plan

    async def analyze(
        self,
        *,
        thread_id: str,
        filename: str,
        content: bytes,
    ) -> LicenseEstate:
        catalog = await self._rate_cards.get()
        parsed, seller_details = parse_customer_document(content, filename)
        return self._analyze_parsed(
            thread_id=thread_id,
            source_file=filename,
            parsed=parsed,
            catalog=catalog,
            seller_details=seller_details,
        )

    async def analyze_parsed(
        self,
        *,
        thread_id: str,
        source_file: str,
        parsed: list[ParsedLicenseRow],
        seller_details: list[SellerProvidedDetail] | None = None,
    ) -> LicenseEstate:
        catalog = await self._rate_cards.get()
        return self._analyze_parsed(
            thread_id=thread_id,
            source_file=source_file,
            parsed=parsed,
            catalog=catalog,
            seller_details=seller_details,
        )

    def _analyze_parsed(
        self,
        *,
        thread_id: str,
        source_file: str,
        parsed: list[ParsedLicenseRow],
        catalog: RateCardCatalog,
        seller_details: list[SellerProvidedDetail] | None = None,
    ) -> LicenseEstate:
        lines = [self._match_row(row, catalog) for row in parsed]
        lines = _aggregate_resolved_lines(lines)
        pending = any(line.match_method == "unresolved" for line in lines)
        now = datetime.now(UTC)
        return LicenseEstate(
            id=thread_id,
            thread_id=thread_id,
            source_file=Path(source_file).name,
            status=(
                EstateStatus.AWAITING_MATCH_CONFIRMATION
                if pending
                else EstateStatus.READY
            ),
            lines=lines,
            rate_card_version=catalog.version,
            seller_details=list(seller_details or []),
            created_at=now,
            updated_at=now,
        )

    def confirm_matches(
        self,
        estate: LicenseEstate,
        selections: Mapping[str, tuple[str, str]],
    ) -> LicenseEstate:
        pending_ids = {line.line_id for line in estate.pending_lines}
        if set(selections) != pending_ids:
            missing = sorted(pending_ids - set(selections))
            unknown = sorted(set(selections) - pending_ids)
            detail = []
            if missing:
                detail.append("missing " + ", ".join(missing))
            if unknown:
                detail.append("unknown " + ", ".join(unknown))
            raise LicenseAnalysisError("Invalid confirmation batch: " + "; ".join(detail))

        updated: list[NormalizedLicenseLine] = []
        for line in estate.lines:
            if line.line_id not in selections:
                updated.append(line)
                continue
            product_id, sku_id = selections[line.line_id]
            candidate = next(
                (
                    item
                    for item in line.candidates
                    if item.product_id.casefold() == product_id.casefold()
                    and item.sku_id.casefold() == sku_id.casefold()
                ),
                None,
            )
            if candidate is None:
                raise LicenseAnalysisError(
                    f"Selection for {line.line_id} is not one of its offered candidates."
                )
            updated.append(
                line.model_copy(
                    update={
                        "product_id": candidate.product_id,
                        "sku_id": candidate.sku_id,
                        "sku_title": candidate.sku_title,
                        "match_confidence": candidate.confidence,
                        "match_method": "seller_confirmed",
                        "candidates": [],
                    }
                )
            )
        return estate.model_copy(
            update={
                "lines": _aggregate_resolved_lines(updated),
                "status": EstateStatus.READY,
                "updated_at": datetime.now(UTC),
            }
        )

    def _match_row(
        self,
        row: ParsedLicenseRow,
        catalog: RateCardCatalog,
    ) -> NormalizedLicenseLine:
        candidates = catalog.candidates(
            row.product_title,
            product_id=row.product_id,
            sku_id=row.sku_id,
            limit=3,
        )
        selected = None
        method: str = "unresolved"
        if len(candidates) == 1 and candidates[0].confidence == 100:
            selected = candidates[0]
            method = "exact"
        # A title-only fuzzy result is always advisory. Even a high score can cross
        # product families (for example, Microsoft 365 E5 versus Dynamics 365), so
        # only an exact catalogue identity is committed without seller clarification.

        return NormalizedLicenseLine(
            line_id=f"L{row.row_number - 1}",
            row_number=row.row_number,
            source_product_title=row.product_title,
            product_id=selected.product_id if selected else None,
            sku_id=selected.sku_id if selected else None,
            sku_title=selected.sku_title if selected else None,
            total_licenses=row.total_licenses,
            expired_licenses=row.expired_licenses,
            assigned_licenses=row.assigned_licenses,
            renewal_quantity=row.renewal_quantity,
            expiration_date=row.expiration_date,
            renewal_date=row.renewal_date,
            term_duration=row.term_duration or self._default_term_duration,
            billing_plan=row.billing_plan or self._default_billing_plan,
            match_confidence=selected.confidence if selected else None,
            match_method=method,  # type: ignore[arg-type]
            candidates=[] if selected else candidates,
        )


def parse_customer_file(content: bytes, filename: str) -> list[ParsedLicenseRow]:
    parsed, _seller_details = parse_customer_document(content, filename)
    return parsed


def parse_customer_document(
    content: bytes,
    filename: str,
) -> tuple[list[ParsedLicenseRow], list[SellerProvidedDetail]]:
    suffix = Path(filename).suffix.casefold()
    if suffix == ".csv":
        rows = list(_csv_rows(content))
    elif suffix in {".xlsx", ".xlsm"}:
        rows = list(_xlsx_rows(content))
    else:
        raise LicenseAnalysisError("Upload a .csv, .xlsx, or .xlsm licence file.")
    return _parse_rows(rows), _seller_details_from_preamble(rows)


def _csv_rows(content: bytes) -> Iterable[Sequence[object]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise LicenseAnalysisError("The customer CSV must be UTF-8 encoded.") from error
    yield from csv.reader(io.StringIO(text))


def _xlsx_rows(content: bytes) -> Iterable[Sequence[object]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        yield from worksheet.iter_rows(values_only=True)
    finally:
        workbook.close()


def _header_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").casefold())


_SELLER_DETAIL_LABELS = {
    "customer": "Customer name",
    "customername": "Customer name",
    "company": "Customer name",
    "companyname": "Customer name",
    "accountname": "Customer name",
    "customerreference": "Customer reference",
    "referencenumber": "Customer reference",
    "referenceid": "Customer reference",
    "opportunity": "Opportunity",
    "opportunityid": "Opportunity",
    "opportunityname": "Opportunity",
    "proposalname": "Proposal name",
    "proposaltitle": "Proposal name",
    "sellernote": "Seller note",
    "notes": "Seller note",
}


def _seller_details_from_preamble(
    rows: Sequence[Sequence[object]],
) -> list[SellerProvidedDetail]:
    """Capture explicit key/value context above the tabular header, if present."""

    details: dict[str, SellerProvidedDetail] = {}
    for candidate in rows[:25]:
        fields = {_HEADER_ALIASES.get(_header_key(value)) for value in candidate}
        if _REQUIRED.issubset(fields):
            break
        populated = [str(value).strip() for value in candidate if value not in (None, "")]
        if len(populated) < 2:
            continue
        label = _SELLER_DETAIL_LABELS.get(_header_key(populated[0]))
        if label is None:
            continue
        value = " ".join(populated[1].split())[:500]
        if value:
            details[label.casefold()] = SellerProvidedDetail(label=label, value=value)
        if len(details) >= 12:
            break
    return list(details.values())


def _parse_rows(rows: Iterable[Sequence[object]]) -> list[ParsedLicenseRow]:
    materialized = list(rows)
    if not materialized:
        raise LicenseAnalysisError("The uploaded licence file is empty.")

    header_index = -1
    raw_headers: Sequence[object] = ()
    for index, candidate_headers in enumerate(materialized[:25]):
        fields = {
            _HEADER_ALIASES.get(_header_key(header)) for header in candidate_headers
        }
        if _REQUIRED.issubset(fields):
            header_index = index
            raw_headers = candidate_headers
            break
    if header_index < 0:
        raise LicenseAnalysisError(
            "Could not find a header row containing SKU/Product Name and Quantity."
        )

    columns: dict[int, str] = {}
    for index, header in enumerate(raw_headers):
        field = _HEADER_ALIASES.get(_header_key(header))
        if field:
            columns[index] = field
    missing = sorted(_REQUIRED - set(columns.values()))
    if missing:
        raise LicenseAnalysisError(
            "Missing required columns: " + ", ".join(missing)
        )

    parsed: list[ParsedLicenseRow] = []
    for row_number, values in enumerate(
        materialized[header_index + 1 :],
        start=header_index + 2,
    ):
        record = {
            field: values[index] if index < len(values) else None
            for index, field in columns.items()
        }
        if not any(value not in (None, "") for value in record.values()):
            continue
        title = str(record.get("product_title") or "").strip()
        if not title:
            raise LicenseAnalysisError(f"Row {row_number}: Product Title is empty.")
        total = _quantity(record.get("total_licenses"), row_number, "Total Licenses")
        expired = _optional_quantity(
            record.get("expired_licenses"), row_number, "Expired Licenses"
        )
        assigned = _optional_quantity(
            record.get("assigned_licenses"), row_number, "Assigned Licenses"
        )
        if expired > total:
            raise LicenseAnalysisError(
                f"Row {row_number}: Expired Licenses cannot exceed Total Licenses."
            )
        if assigned > total:
            raise LicenseAnalysisError(
                f"Row {row_number}: Assigned Licenses cannot exceed Total Licenses."
            )
        parsed.append(
            ParsedLicenseRow(
                row_number=row_number,
                product_title=title,
                product_id=_optional_text(record.get("product_id")),
                sku_id=_optional_text(record.get("sku_id")),
                total_licenses=total,
                expired_licenses=expired,
                assigned_licenses=assigned,
                renewal_quantity=total - expired,
                expiration_date=_date(record.get("expiration_date"), row_number),
                renewal_date=_date(record.get("renewal_date"), row_number),
                term_duration=_optional_text(record.get("term_duration")),
                billing_plan=_optional_text(record.get("billing_plan")),
            )
        )
    if not parsed:
        raise LicenseAnalysisError("The licence file contains no data rows.")
    return parsed


def _quantity(value: object, row_number: int, name: str) -> int:
    if isinstance(value, bool):
        raise LicenseAnalysisError(f"Row {row_number}: {name} must be an integer.")
    try:
        if isinstance(value, float) and not value.is_integer():
            raise ValueError
        result = int(str(value).strip()) if not isinstance(value, (int, float)) else int(value)
    except (TypeError, ValueError) as error:
        raise LicenseAnalysisError(
            f"Row {row_number}: {name} must be a whole number."
        ) from error
    if result < 0:
        raise LicenseAnalysisError(f"Row {row_number}: {name} cannot be negative.")
    return result


def _optional_quantity(value: object, row_number: int, name: str) -> int:
    if value in (None, ""):
        return 0
    return _quantity(value, row_number, name)


def _optional_text(value: object) -> str | None:
    text = str(value or "").strip()
    return text or None


def _date(value: object, row_number: int) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    raise LicenseAnalysisError(
        f"Row {row_number}: date {text!r} is not in a supported format."
    )


def _aggregate_resolved_lines(
    lines: list[NormalizedLicenseLine],
) -> list[NormalizedLicenseLine]:
    grouped: dict[
        tuple[str, str, date | None, date | None],
        tuple[NormalizedLicenseLine, int],
    ] = {}
    result: list[NormalizedLicenseLine] = []
    for line in lines:
        if not line.product_id or not line.sku_id:
            result.append(line)
            continue
        key = (line.product_id, line.sku_id, line.expiration_date, line.renewal_date)
        stored = grouped.get(key)
        if stored is None:
            grouped[key] = (line, len(result))
            result.append(line)
            continue
        current, result_index = stored
        merged = current.model_copy(
            update={
                "source_product_title": (
                    current.source_product_title
                    if current.source_product_title == line.source_product_title
                    else current.source_product_title + "; " + line.source_product_title
                ),
                "total_licenses": current.total_licenses + line.total_licenses,
                "expired_licenses": current.expired_licenses + line.expired_licenses,
                "assigned_licenses": current.assigned_licenses + line.assigned_licenses,
                "renewal_quantity": current.renewal_quantity + line.renewal_quantity,
                "term_duration": current.term_duration or line.term_duration,
                "billing_plan": current.billing_plan or line.billing_plan,
            }
        )
        grouped[key] = (merged, result_index)
        result[result_index] = merged
    return result
