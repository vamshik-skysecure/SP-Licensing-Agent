from __future__ import annotations

import asyncio
import csv
import hashlib
import io
import re
import unicodedata
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from time import monotonic
from typing import Protocol

from rapidfuzz import fuzz

from .models import RateCardItem, SkuMatchCandidate


class RateCardError(ValueError):
    pass


@dataclass(frozen=True)
class RateCardPayload:
    content: bytes
    filename: str
    version: str


class RateCardSource(Protocol):
    async def fetch(self) -> RateCardPayload: ...

    async def close(self) -> None: ...


class LocalRateCardSource:
    def __init__(self, path: Path) -> None:
        self._path = path

    async def fetch(self) -> RateCardPayload:
        return await asyncio.to_thread(self._read)

    def _read(self) -> RateCardPayload:
        try:
            content = self._path.read_bytes()
        except PermissionError as error:
            raise RateCardError(
                f"The local rate-card workbook {self._path} is locked. Close it in "
                "Excel/OneDrive or use the Azure Blob backend."
            ) from error
        stat = self._path.stat()
        version = f"local:{stat.st_mtime_ns}:{stat.st_size}"
        return RateCardPayload(content, self._path.name, version)

    async def close(self) -> None:
        return None


class AzureBlobRateCardSource:
    """Reads the maintained workbook directly from Blob using MI or a dev connection string."""

    def __init__(
        self,
        *,
        container_name: str,
        blob_name: str,
        account_url: str | None = None,
        connection_string: str | None = None,
    ) -> None:
        if not (connection_string or account_url):
            raise ValueError("An account URL or connection string is required.")

        from azure.storage.blob.aio import BlobServiceClient

        self._credential = None
        if connection_string:
            self._service = BlobServiceClient.from_connection_string(connection_string)
        else:
            from azure.identity.aio import DefaultAzureCredential

            self._credential = DefaultAzureCredential()
            self._service = BlobServiceClient(
                account_url=account_url,
                credential=self._credential,
            )
        self._blob = self._service.get_blob_client(container_name, blob_name)
        self._blob_name = blob_name

    async def fetch(self) -> RateCardPayload:
        properties = await self._blob.get_blob_properties()
        stream = await self._blob.download_blob()
        content = await stream.readall()
        version = str(properties.etag or properties.last_modified or len(content))
        return RateCardPayload(content, self._blob_name, version)

    async def close(self) -> None:
        await self._service.close()
        if self._credential is not None:
            await self._credential.close()


def normalize_product_title(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value).casefold()
    normalized = re.sub(r"[^\w]+", " ", normalized, flags=re.UNICODE)
    normalized = " ".join(normalized.split())
    aliases = (
        # Sellers commonly use ME3/ME5/ME7 as shorthand for the Microsoft 365
        # Enterprise suites. Resolve those terms before fuzzy matching so a
        # quantity such as "ME7 1 qty" cannot make an unrelated "1 year" SKU
        # look like the closest catalogue result.
        (r"\bme([357])\b", r"microsoft 365 e\1"),
        (r"\bm365\b", "microsoft 365"),
        (r"\bo365\b", "office 365"),
        (r"\bems\b", "enterprise mobility security"),
        (r"\bplan one\b", "plan 1"),
        (r"\bplan two\b", "plan 2"),
    )
    for pattern, replacement in aliases:
        normalized = re.sub(pattern, replacement, normalized)
    return normalized


_GENERIC_SKU_TOKENS = {
    "license",
    "licenses",
    "licence",
    "licences",
    "plan",
    "plans",
    "product",
    "sku",
    "subscription",
    "subscriptions",
}
_VARIANT_MARKERS = (
    "education",
    "student",
    "faculty",
    "non profit",
    "nonprofit",
    "government",
    "charity",
    "three year",
    "3 year",
    "36 month",
    "unattended",
    "student use benefit",
    "no teams",
    "add on",
)
_REQUIRED_QUERY_QUALIFIERS = frozenset({"premium", "studio"})
_DEFENDER_WORKLOAD_MARKERS = (
    ("vulnerability management", "vulnerability-management"),
    ("office 365", "office-365"),
    ("cloud apps", "cloud-apps"),
    ("endpoint", "endpoint"),
    ("identity", "identity"),
    ("iot", "iot"),
    ("business", "business"),
    ("purview", "purview"),
)
_FAMILY_MARKERS = (
    # Specific product families must precede the suite names embedded in their
    # titles (for example, Defender for Office 365 and Microsoft 365 Copilot).
    ("microsoft defender", "defender"),
    ("defender", "defender"),
    ("copilot", "copilot"),
    ("dynamics 365", "dynamics-365"),
    ("enterprise mobility security", "enterprise-mobility-security"),
    ("power bi", "power-bi"),
    ("power apps", "power-platform"),
    ("power automate", "power-platform"),
    ("microsoft 365", "microsoft-365"),
    ("office 365", "office-365"),
    ("teams", "teams"),
    ("visio", "visio"),
    ("project", "project"),
    ("intune", "intune"),
    ("entra", "entra"),
    ("windows", "windows"),
)
_AMBIGUOUS_TIER_FAMILY_PRIORITY = {
    "microsoft-365": 0,
    "office-365": 1,
    "enterprise-mobility-security": 2,
    "defender": 3,
    "dynamics-365": 4,
}


def _tier_tokens(normalized: str) -> set[str]:
    return set(re.findall(r"\b(?:a|e|f)\d+\b", normalized))


def _plan_tokens(normalized: str) -> set[str]:
    """Canonicalize spoken plan wording without changing the display title.

    Microsoft uses both ``Plan 2`` and ``P2`` across catalogue families. ``F2`` is
    a separate frontline licence and must never satisfy a seller request for Plan 2.
    """

    result = {
        f"{prefix}{number}"
        for prefix, number in re.findall(r"\b([pf])\s*([0-9]+)\b", normalized)
    }
    result.update(
        f"p{number}" for number in re.findall(r"\bplan\s+([0-9]+)\b", normalized)
    )
    return result


def _defender_workloads(normalized: str) -> set[str]:
    if "defender" not in normalized:
        return set()
    return {
        workload
        for marker, workload in _DEFENDER_WORKLOAD_MARKERS
        if marker in normalized
    }


def _product_family_key(normalized: str) -> str | None:
    for marker, family in _FAMILY_MARKERS:
        if marker in normalized:
            return family
    return None


def _unrequested_variant_count(query: str, title: str) -> int:
    return sum(marker in title and marker not in query for marker in _VARIANT_MARKERS)


def _base_suite_penalty(title: str) -> int:
    """Prefer a suite itself over similarly named add-ons for a tier-only query."""

    tier = next(iter(_tier_tokens(title)), "")
    family = _product_family_key(title)
    if not tier or family not in {
        "microsoft-365",
        "office-365",
        "enterprise-mobility-security",
    }:
        return 1
    permitted = {
        "microsoft",
        "office",
        "365",
        "enterprise",
        "mobility",
        "security",
        tier,
        "without",
        "audio",
        "conferencing",
        "no",
        "teams",
    }
    return int(bool(set(title.split()) - permitted))


@dataclass(frozen=True)
class SkuIdentity:
    product_id: str
    sku_id: str
    sku_title: str


class RateCardCatalog:
    def __init__(self, items: Sequence[RateCardItem], version: str) -> None:
        if not items:
            raise RateCardError("The Outcome Sheet contains no priceable rows.")
        self.items = tuple(items)
        self.version = version
        identities: dict[tuple[str, str, str], SkuIdentity] = {}
        for item in items:
            key = (item.product_id, item.sku_id, item.sku_title)
            identities.setdefault(key, SkuIdentity(*key))
        self.identities = tuple(identities.values())
        self._normalized_titles = {
            index: normalize_product_title(identity.sku_title)
            for index, identity in enumerate(self.identities)
        }
        self._items_by_identity: dict[tuple[str, str, str], list[RateCardItem]] = {}
        for item in items:
            key = (item.product_id, item.sku_id, item.sku_title)
            self._items_by_identity.setdefault(key, []).append(item)

    def candidates(
        self,
        query: str,
        *,
        product_id: str | None = None,
        sku_id: str | None = None,
        limit: int | None = 3,
    ) -> list[SkuMatchCandidate]:
        if (
            product_id
            and sku_id
            and not _is_placeholder_identifier(product_id)
            and not _is_placeholder_identifier(sku_id)
        ):
            exact_ids = [
                identity
                for identity in self.identities
                if identity.product_id.casefold() == product_id.casefold()
                and identity.sku_id.casefold() == sku_id.casefold()
            ]
            if exact_ids:
                normalized_query = normalize_product_title(query)
                exact_title = [
                    identity
                    for identity in exact_ids
                    if normalize_product_title(identity.sku_title) == normalized_query
                ]
                if exact_title:
                    matches = [
                        self._candidate(identity, 100.0)
                        for identity in exact_title
                    ]
                    return matches if limit is None else matches[:limit]
                if len(exact_ids) == 1:
                    return [self._candidate(exact_ids[0], 100.0)]
                best = max(
                    exact_ids,
                    key=lambda identity: fuzz.WRatio(
                        normalized_query,
                        normalize_product_title(identity.sku_title),
                    ),
                )
                return [self._candidate(best, 100.0)]

        normalized = normalize_product_title(query)
        exact = [
            self.identities[index]
            for index, title in self._normalized_titles.items()
            if title == normalized
        ]
        if exact:
            # The same display title can exist under distinct Microsoft Product/SKU
            # identities. Do not collapse those identities or silently select one;
            # the seller must see and confirm every maintained catalogue choice.
            matches = [
                self._candidate(identity, 100.0)
                for identity in sorted(exact, key=self._identity_preference)
            ]
            if limit is None:
                return matches
            # A finite limit is the legacy internal preview API. Keep one preferred
            # identity there; all seller-facing analysis explicitly requests None.
            return matches[:1]

        query_tokens = set(normalized.split())
        required_qualifiers = query_tokens & _REQUIRED_QUERY_QUALIFIERS
        tiers = _tier_tokens(normalized)
        requested_plans = _plan_tokens(normalized)
        requested_defender_workloads = _defender_workloads(normalized)
        query_family = _product_family_key(normalized)
        pool = list(self.identities)
        if tiers:
            tier_pool = [
                identity
                for identity in pool
                if tiers.issubset(_tier_tokens(normalize_product_title(identity.sku_title)))
            ]
            if not tier_pool:
                return []
            pool = tier_pool
        if query_family:
            family_pool = [
                identity
                for identity in pool
                if _product_family_key(normalize_product_title(identity.sku_title))
                == query_family
            ]
            if not family_pool:
                return []
            pool = family_pool

        ranked: list[tuple[tuple[object, ...], SkuIdentity, float]] = []
        for identity in pool:
            title = normalize_product_title(identity.sku_title)
            title_tokens = set(title.split())
            if requested_plans and not requested_plans.issubset(_plan_tokens(title)):
                continue
            if requested_defender_workloads and not requested_defender_workloads.issubset(
                _defender_workloads(title)
            ):
                continue
            if required_qualifiers and not required_qualifiers.issubset(title_tokens):
                continue
            informative_query = query_tokens - _GENERIC_SKU_TOKENS
            overlap = len(informative_query & title_tokens)
            if informative_query and overlap == 0:
                continue
            score = float(fuzz.WRatio(normalized, title))
            if informative_query:
                score += 8.0 * overlap / len(informative_query)
            score -= 6.0 * _unrequested_variant_count(normalized, title)
            score = max(0.0, min(99.0, score))
            if score < 55.0:
                continue
            ranked.append(
                (
                    self._candidate_rank(
                        identity,
                        normalized_query=normalized,
                        query_family=query_family,
                        has_tier=bool(tiers),
                        score=score,
                    ),
                    identity,
                    score,
                )
            )
        ranked.sort(key=lambda item: item[0])

        result: list[SkuMatchCandidate] = []
        seen_titles: set[str] = set()
        seen_families: set[str] = set()
        family_preview = limit is not None and bool(tiers) and query_family is None
        for _rank, identity, score in ranked:
            # Identities are already unique by ProductId + SkuId + title. Unlimited
            # seller-facing lookups preserve every identity; finite internal previews
            # retain their historical title/family diversification.
            title_key = normalize_product_title(identity.sku_title)
            family_key = _product_family_key(title_key) or title_key
            if limit is not None and title_key in seen_titles:
                continue
            if family_preview and family_key in seen_families:
                continue
            seen_titles.add(title_key)
            seen_families.add(family_key)
            result.append(self._candidate(identity, score))
            if limit is not None and len(result) >= limit:
                break
        return result

    def _preferred_identity(self, identities: Sequence[SkuIdentity]) -> SkuIdentity:
        return min(identities, key=self._identity_preference)

    def _identity_preference(self, identity: SkuIdentity) -> tuple[object, ...]:
        rows = self._items_by_identity.get(
            (identity.product_id, identity.sku_id, identity.sku_title),
            [],
        )
        annual = any(
            row.term_duration.casefold() == "p1y"
            and row.billing_plan.casefold() == "annual"
            for row in rows
        )
        commercial = any(
            (row.segment or "").casefold() == "commercial"
            and row.term_duration.casefold() == "p1y"
            and row.billing_plan.casefold() == "annual"
            for row in rows
        )
        priceable = any(
            (
                row.distributor_price > 0
                or row.marketplace_price > 0
                or row.initial_quote_with_promo_available
                or row.initial_quote_without_promo_available
            )
            and row.term_duration.casefold() == "p1y"
            and row.billing_plan.casefold() == "annual"
            for row in rows
        )
        return (
            not commercial,
            not annual,
            not priceable,
            _unrequested_variant_count("", normalize_product_title(identity.sku_title)),
            identity.sku_id.casefold(),
        )

    def _candidate_rank(
        self,
        identity: SkuIdentity,
        *,
        normalized_query: str,
        query_family: str | None,
        has_tier: bool,
        score: float,
    ) -> tuple[object, ...]:
        title = normalize_product_title(identity.sku_title)
        family = _product_family_key(title)
        family_priority = (
            _AMBIGUOUS_TIER_FAMILY_PRIORITY.get(family or "", 99)
            if has_tier and query_family is None
            else 0
        )
        return (
            family_priority,
            _unrequested_variant_count(normalized_query, title),
            _base_suite_penalty(title),
            *self._identity_preference(identity),
            -score,
            len(title),
            title,
        )

    def price_rows(
        self,
        *,
        product_id: str,
        sku_id: str,
        sku_title: str | None = None,
        term_duration: str | None = None,
        billing_plan: str | None = None,
        segment: str | None = None,
    ) -> list[RateCardItem]:
        rows = [
            item
            for item in self.items
            if item.product_id.casefold() == product_id.casefold()
            and item.sku_id.casefold() == sku_id.casefold()
        ]
        if sku_title:
            normalized = normalize_product_title(sku_title)
            title_rows = [
                item
                for item in rows
                if normalize_product_title(item.sku_title) == normalized
            ]
            if title_rows:
                rows = title_rows
        if term_duration:
            rows = [
                item
                for item in rows
                if item.term_duration.casefold() == term_duration.casefold()
            ]
        if billing_plan:
            rows = [
                item
                for item in rows
                if item.billing_plan.casefold() == billing_plan.casefold()
            ]
        if segment:
            segment_rows = [
                item
                for item in rows
                if item.segment and item.segment.casefold() == segment.casefold()
            ]
            if segment_rows:
                rows = segment_rows
        return rows

    def higher_tier_candidates(
        self,
        current_title: str,
        *,
        limit: int = 3,
    ) -> list[SkuMatchCandidate]:
        """Return same-family higher-tier catalogue options for seller review only."""

        normalized = normalize_product_title(current_title)
        family = _product_family_key(normalized)
        current_tiers = _tier_tokens(normalized)
        if family is None or len(current_tiers) != 1:
            return []
        current_tier = next(iter(current_tiers))
        tier_match = re.fullmatch(r"([aef])(\d+)", current_tier)
        if tier_match is None:
            return []
        tier_prefix, current_level_text = tier_match.groups()
        current_level = int(current_level_text)
        by_tier: dict[int, list[SkuIdentity]] = {}
        for identity in self.identities:
            title = normalize_product_title(identity.sku_title)
            if _product_family_key(title) != family:
                continue
            tiers = _tier_tokens(title)
            if len(tiers) != 1:
                continue
            match = re.fullmatch(r"([aef])(\d+)", next(iter(tiers)))
            if match is None or match.group(1) != tier_prefix:
                continue
            level = int(match.group(2))
            if level <= current_level:
                continue
            by_tier.setdefault(level, []).append(identity)

        result: list[SkuMatchCandidate] = []
        for level in sorted(by_tier):
            identity = min(
                by_tier[level],
                key=lambda item: (
                    _unrequested_variant_count(
                        normalized,
                        normalize_product_title(item.sku_title),
                    ),
                    _base_suite_penalty(normalize_product_title(item.sku_title)),
                    *self._identity_preference(item),
                    len(item.sku_title),
                ),
            )
            result.append(self._candidate(identity, 99.0))
            if len(result) >= limit:
                break
        return result

    @staticmethod
    def _candidate(identity: SkuIdentity, confidence: float) -> SkuMatchCandidate:
        return SkuMatchCandidate(
            product_id=identity.product_id,
            sku_id=identity.sku_id,
            sku_title=identity.sku_title,
            confidence=confidence,
        )


class RateCardProvider:
    def __init__(
        self,
        source: RateCardSource,
        *,
        sheet_name: str = "Outcome Sheet",
        refresh_seconds: float = 300,
    ) -> None:
        self._source = source
        self._sheet_name = sheet_name
        self._refresh_seconds = refresh_seconds
        self._catalog: RateCardCatalog | None = None
        self._loaded_at = 0.0
        self._lock = asyncio.Lock()

    async def get(self, *, force: bool = False) -> RateCardCatalog:
        if (
            not force
            and self._catalog is not None
            and monotonic() - self._loaded_at < self._refresh_seconds
        ):
            return self._catalog
        async with self._lock:
            if (
                not force
                and self._catalog is not None
                and monotonic() - self._loaded_at < self._refresh_seconds
            ):
                return self._catalog
            payload = await self._source.fetch()
            items = await asyncio.to_thread(
                parse_rate_card,
                payload.content,
                payload.filename,
                self._sheet_name,
            )
            digest = hashlib.sha256(payload.content).hexdigest()[:16]
            self._catalog = RateCardCatalog(
                items,
                version=f"{payload.version}:{digest}",
            )
            self._loaded_at = monotonic()
            return self._catalog

    async def close(self) -> None:
        await self._source.close()


_HEADERS = {
    "productid": "product_id",
    "skuid": "sku_id",
    "skutitle": "sku_title",
    "contracttype": "contract_type",
    "termduration": "term_duration",
    "billingplan": "billing_plan",
    "segment": "segment",
    "erp": "erp_price",
    "erpprice": "erp_price",
    "catalogueprice": "catalogue_price",
    "unitpricecatalogue": "catalogue_price",
    "promoname": "promo_name",
    "promo": "promo_percentage",
    "promoifnewtoms": "promo_percentage",
    "customereligibility": "customer_eligibility",
    "newtomicrosoftrequired": "new_to_microsoft_required",
    "minimumseats": "minimum_seats",
    "maximumseats": "maximum_seats",
    "geography": "geography",
    "nettoms": "net_to_ms",
    "expectedpartnerpricingwithpromo": "partner_price_with_promo",
    "expectedpartnerpricingwithoutpromo": "partner_price_without_promo",
    "distributorlandingprice": "distributor_landing_price",
    "partnerlandingpricefromdistributor": "partner_landing_price",
    "partnerscostofproduct": "partner_cost",
    "partnerbestoffer": "partner_best_offer",
    "priceonmarketplace": "marketplace_price",
    # V6.0 currently contains the misspelled "Expectec" header. Support both
    # spellings so a corrected business workbook does not require a code change.
    "expectecdistipricetoskysecure": "distributor_price",
    "expecteddistipricetoskysecure": "distributor_price",
    "initialquotewithpromo": "initial_quote_with_promo",
    "initialquotewithoutpromo": "initial_quote_without_promo",
}

_BASE_REQUIRED = {
    "product_id",
    "sku_id",
    "sku_title",
    "term_duration",
    "billing_plan",
}


def parse_rate_card(
    content: bytes,
    filename: str,
    sheet_name: str = "Outcome Sheet",
) -> list[RateCardItem]:
    suffix = Path(filename).suffix.casefold()
    if suffix == ".xlsx":
        rows = _xlsx_rows(content, sheet_name)
    elif suffix == ".csv":
        rows = _csv_rows(content)
    else:
        raise RateCardError("The rate card must be an .xlsx workbook or .csv file.")
    return _parse_rows(rows)


def _xlsx_rows(content: bytes, sheet_name: str) -> Iterable[Sequence[object]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise RateCardError(f"Workbook sheet {sheet_name!r} was not found.")
        yield from workbook[sheet_name].iter_rows(values_only=True)
    finally:
        workbook.close()


def _csv_rows(content: bytes) -> Iterable[Sequence[object]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise RateCardError("The rate-card CSV must be UTF-8 encoded.") from error
    yield from csv.reader(io.StringIO(text))


def _header_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").casefold())


def _is_placeholder_identifier(value: str) -> bool:
    return value.strip().casefold() in {"0", "-", "n/a", "na", "none", "unmapped"}


def _parse_rows(rows: Iterable[Sequence[object]]) -> list[RateCardItem]:
    iterator = iter(rows)
    columns: dict[int, str] = {}
    header_row_number = 0
    # Business workbooks commonly place a title or blank row above the real
    # headers. Locate the first complete catalogue header within a bounded area.
    for candidate_row_number, raw_headers in enumerate(iterator, start=1):
        candidate_columns = {
            index: field
            for index, header in enumerate(raw_headers)
            if (field := _HEADERS.get(_header_key(header)))
        }
        if _BASE_REQUIRED.issubset(candidate_columns.values()):
            columns = candidate_columns
            header_row_number = candidate_row_number
            break
        if candidate_row_number >= 25:
            break
    if not columns:
        raise RateCardError(
            "A complete rate-card header was not found in the first 25 rows."
        )
    available_fields = set(columns.values())
    missing = sorted(_BASE_REQUIRED - available_fields)
    if missing:
        raise RateCardError("Missing rate-card columns: " + ", ".join(missing))
    has_legacy_quotes = {
        "initial_quote_with_promo",
        "initial_quote_without_promo",
    }.issubset(available_fields)
    has_final_offer = "partner_best_offer" in available_fields
    has_marketplace_price = "marketplace_price" in available_fields
    has_distributor_price = "distributor_price" in available_fields
    if not any(
        (
            has_legacy_quotes,
            has_final_offer,
            has_marketplace_price,
            has_distributor_price,
        )
    ):
        raise RateCardError(
            "Missing rate-card price columns: provide a supported commercial-price field."
        )

    items: list[RateCardItem] = []
    for row_number, values in enumerate(iterator, start=header_row_number + 1):
        record = {
            field: values[index] if index < len(values) else None
            for index, field in columns.items()
        }
        if not any(value not in (None, "") for value in record.values()):
            continue
        # V6 includes rows without Microsoft ProductId/SkuId. Retain them as
        # name-only catalogue entries; an empty identifier is never invented or
        # displayed. Duplicate same-name commercial prices remain ambiguous and
        # are rejected by ScenarioEngine rather than selected silently.
        for field in ("product_id", "sku_id"):
            raw_value = record.get(field)
            record[field] = "" if raw_value in (None, "") else str(raw_value).strip()
        for field in ("sku_title", "billing_plan"):
            raw_value = record.get(field)
            record[field] = "" if raw_value in (None, "") else str(raw_value).strip()
            if not record[field]:
                raise RateCardError(f"Rate-card row {row_number} has an empty {field}.")
        raw_term = record.get("term_duration")
        record["term_duration"] = (
            str(raw_term).strip()
            if raw_term not in (None, "")
            else (
                "Perpetual"
                if record["billing_plan"].casefold() in {"onetime", "one time"}
                else ""
            )
        )
        if not record["term_duration"]:
            raise RateCardError(
                f"Rate-card row {row_number} has an empty term_duration."
            )

        # Microsoft SKU V5.0 makes the maintained Final Output Sheet authoritative.
        # Its direct seller offer is the quoted commercial value. Promotional rows are
        # deliberately not exposed as a non-promotional price: the seller must first
        # confirm eligibility, while standard rows remain usable in either mode.
        if has_final_offer and not has_legacy_quotes:
            offer = record.get("partner_best_offer")
            normalized_offer = _decimal(offer, "partner_best_offer", row_number)
            promo_percentage = _decimal(
                record.get("promo_percentage"), "promo_percentage", row_number
            )
            promo_name = str(record.get("promo_name") or "").strip()
            is_promotional = promo_percentage > 0 or promo_name not in {"", "0"}
            offer_is_available = offer not in (None, "") and normalized_offer > 0
            record["initial_quote_with_promo"] = (
                offer if is_promotional and offer_is_available else None
            )
            record["initial_quote_without_promo"] = (
                offer if not is_promotional and offer_is_available else None
            )

        record["initial_quote_with_promo_available"] = record.get(
            "initial_quote_with_promo"
        ) not in (None, "")
        record["initial_quote_without_promo_available"] = record.get(
            "initial_quote_without_promo"
        ) not in (None, "")
        for field in (
            "erp_price",
            "catalogue_price",
            "promo_percentage",
            "net_to_ms",
            "partner_price_with_promo",
            "partner_price_without_promo",
            "distributor_landing_price",
            "partner_landing_price",
            "partner_cost",
            "partner_best_offer",
            "marketplace_price",
            "distributor_price",
            "initial_quote_with_promo",
            "initial_quote_without_promo",
        ):
            record[field] = _decimal(record.get(field), field, row_number)
        for field in (
            "contract_type",
            "segment",
            "promo_name",
            "customer_eligibility",
            "new_to_microsoft_required",
            "geography",
        ):
            value = record.get(field)
            normalized = str(value).strip() if value not in (None, "", 0, "0") else None
            record[field] = normalized
        for field in ("minimum_seats", "maximum_seats"):
            record[field] = _optional_int(record.get(field), field, row_number)
        items.append(RateCardItem(**record, source_row_number=row_number))
    return items


def _decimal(value: object, field: str, row_number: int) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        result = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError) as error:
        raise RateCardError(
            f"Rate-card row {row_number} has an invalid {field}."
        ) from error
    if result < 0:
        raise RateCardError(f"Rate-card row {row_number} has a negative {field}.")
    # Excel formulas commonly surface harmless binary-float tails such as
    # 1746.3600000000001. Normalize at the ingestion boundary while retaining
    # more precision than the two-decimal commercial output.
    return result.quantize(Decimal("0.000001"))


def _optional_int(value: object, field: str, row_number: int) -> int | None:
    if value in (None, ""):
        return None
    try:
        result = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError) as error:
        raise RateCardError(
            f"Rate-card row {row_number} has an invalid {field}."
        ) from error
    if result < 0 or result != result.to_integral_value():
        raise RateCardError(
            f"Rate-card row {row_number} has an invalid {field}."
        )
    return int(result)
